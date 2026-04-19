from __future__ import annotations

import json
import math
import re
import statistics
import subprocess
import unicodedata
import zipfile
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image


SOURCE_ROOT = Path(r"E:\info de campo")
CONSULTANCY_ROOT = Path(r"E:\CONSULTORIA PARA LA ELABORACION DE UN PROCEDIMIENTO TECNICO PARA LA REGULARIZACION DE AREAS SENSIBLES")
REPO_ROOT = Path(__file__).resolve().parents[1]
OUTPUT_PATH = REPO_ROOT / "public-data" / "field" / "field_evidence.json"
WORK_TMP = REPO_ROOT / "tmp" / "field-consultancy"
Image.MAX_IMAGE_PIXELS = None

OGRINFO_CANDIDATES = [
    Path(r"C:\Program Files\QGIS 3.40.10\bin\ogrinfo.exe"),
    Path(r"C:\Program Files\QGIS 3.40.10\apps\bin\ogrinfo.exe"),
]
OGR2OGR_CANDIDATES = [
    Path(r"C:\Program Files\QGIS 3.40.10\bin\ogr2ogr.exe"),
    Path(r"C:\Program Files\QGIS 3.40.10\apps\bin\ogr2ogr.exe"),
]

CONSULTANCY_SUBPRODUCT_GDB = CONSULTANCY_ROOT / "GDB_SUBPRODUCTOS_AREAS_SENSIBLES" / "GDB_SUBPRODUCTOS_AREAS_SENSIBLES.gdb"
CONSULTANCY_BASE_GDB = CONSULTANCY_ROOT / "GDB_AREAS_SENSIBLES" / "GDB_AREAS_SENSIBLES_MEJIA.gdb"
CONSULTANCY_TYPOLOGY_GDB = (
    CONSULTANCY_ROOT
    / "GDB_TIPOLOGIA_AREAS_SENSIBLES"
    / "TIPOLOGIA_AREAS_SENSIBLES_MEJIA"
    / "TIPOLOGIA_AREAS_SENSIBLES_MEJIA.gdb"
)
CONSULTANCY_PROJECTS = [
    CONSULTANCY_ROOT / "GDB_SUBPRODUCTOS_AREAS_SENSIBLES" / "GDB_ACCID_GEO.aprx",
    CONSULTANCY_ROOT / "GDB_TIPOLOGIA_AREAS_SENSIBLES" / "TIPOLOGIA_AREAS_SENSIBLES_MEJIA" / "TIPOL_MEJIA.aprx",
]
CONSULTANCY_GDBS = [
    CONSULTANCY_ROOT / "CATALOGO_DE_OBJETOS_MEJIA" / "GEODATABASE_MEJIA.gdb",
    CONSULTANCY_BASE_GDB,
    CONSULTANCY_SUBPRODUCT_GDB,
    CONSULTANCY_TYPOLOGY_GDB,
]
CONSULTANCY_SUBPRODUCT_THEMES = [
    {
        "id": "area-proteccion",
        "layer": "area_protec",
        "label": "Área de protección sensible",
        "group": "Protección mínima",
        "tone": "high",
        "description": "Polígonos derivados para resguardo y regularización en áreas sensibles.",
    },
    {
        "id": "vegetacion-riberena",
        "layer": "vege_ribere",
        "label": "Vegetación ribereña",
        "group": "Cobertura ribereña",
        "tone": "resilience",
        "description": "Cobertura ribereña sensible asociada a cauces, quebradas y franjas de protección.",
    },
    {
        "id": "quebradas-sensibles",
        "layer": "quebrada",
        "label": "Quebradas sensibles",
        "group": "Hidrografía sensible",
        "tone": "service",
        "description": "Trazos de quebrada priorizados en la consultoría para control y regularización.",
    },
    {
        "id": "red-hidrica",
        "layer": "stream_q",
        "label": "Red hídrica priorizada",
        "group": "Hidrografía sensible",
        "tone": "service",
        "description": "Cauces principales empleados en el modelamiento de áreas sensibles.",
    },
    {
        "id": "bordes-superiores",
        "layer": "bord_s_que",
        "label": "Bordes superiores de quebrada",
        "group": "Protección mínima",
        "tone": "mid",
        "description": "Líneas de borde superior empleadas para vigilar ocupación en quebradas abiertas.",
    },
    {
        "id": "cambio-pendiente",
        "layer": "cam_pendie",
        "label": "Cambio de pendiente",
        "group": "Geomorfología",
        "tone": "mid",
        "description": "Líneas geomorfológicas donde cambia la pendiente y sube la sensibilidad territorial.",
    },
    {
        "id": "variables-intermedias",
        "layer": "VARIABLES_INTERMEDIAS",
        "label": "Variables intermedias",
        "group": "Modelación multicriterio",
        "tone": "base",
        "description": "Polígonos síntesis de variables intermedias usadas en el procedimiento técnico.",
    },
]
CONSULTANCY_BASE_THEMES = [
    {"layer": "AA01_ACEQUIA", "label": "Acequias", "group": "Aguas interiores"},
    {"layer": "AA02_QUEBRADA", "label": "Quebradas", "group": "Aguas interiores"},
    {"layer": "AA03_ALVEOLOS", "label": "Alvéolos", "group": "Aguas interiores"},
    {"layer": "BA01_CAMBIO_DE_PENDIENTE", "label": "Cambio de pendiente", "group": "Geomorfología"},
    {"layer": "BB01_TALUD_NATURAL", "label": "Talud natural", "group": "Accidentes geográficos"},
    {
        "layer": "BB02_BORDE_SUPERIOR_DE_LA_QUEBRADA_ABIERTA",
        "label": "Borde superior de la quebrada abierta",
        "group": "Accidentes geográficos",
    },
    {
        "layer": "BB03_BORDE_SUPERIOR_DE_LA_ACEQUIA",
        "label": "Borde superior de la acequia",
        "group": "Accidentes geográficos",
    },
    {"layer": "CC02_AREA_DE_PROTECCIÓN", "label": "Área de protección", "group": "Demarcación"},
    {"layer": "DA01_VEGETACION_RIBEREÑA", "label": "Vegetación ribereña", "group": "Cobertura"},
]
CONSULTANCY_TYPOLOGY_THEMES = [
    {"layer": "ABIERTA", "label": "Quebrada abierta", "group": "Tipología quebrada"},
    {"layer": "RELLENA", "label": "Quebrada rellena", "group": "Tipología quebrada"},
    {"layer": "EJE", "label": "Eje de quebrada", "group": "Tipología quebrada"},
    {"layer": "NATURAL", "label": "Talud natural", "group": "Tipología talud"},
    {"layer": "ARTIFICIAL", "label": "Talud artificial", "group": "Tipología talud"},
    {"layer": "ALVEO", "label": "Álveo", "group": "Tipología río"},
    {"layer": "RIBERA", "label": "Ribera", "group": "Tipología río"},
    {"layer": "ESTANQUE", "label": "Estanque", "group": "Tipología cuerpo de agua"},
    {"layer": "LAGO", "label": "Lago", "group": "Tipología cuerpo de agua"},
    {"layer": "RESERVORIO", "label": "Reservorio", "group": "Tipología cuerpo de agua"},
    {"layer": "PRESA", "label": "Presa", "group": "Tipología cuerpo de agua"},
    {"layer": "GRANJ_ACU", "label": "Granja acuícola", "group": "Tipología cuerpo de agua"},
    {"layer": "AREA_RELLEN", "label": "Área rellenada", "group": "Tipología auxiliar"},
    {"layer": "AREA_PROTEC", "label": "Área de protección", "group": "Tipología auxiliar"},
]

MONTH_CODES = {
    "ENE": 0,
    "FEB": 1,
    "MAR": 2,
    "ABR": 3,
    "MAY": 4,
    "JUN": 5,
    "JUL": 6,
    "AGO": 7,
    "SEP": 8,
    "OCT": 9,
    "NOV": 10,
    "DIC": 11,
}
MONTH_NAMES = [
    "Enero",
    "Febrero",
    "Marzo",
    "Abril",
    "Mayo",
    "Junio",
    "Julio",
    "Agosto",
    "Septiembre",
    "Octubre",
    "Noviembre",
    "Diciembre",
]


def discover_binary(candidates: list[Path]) -> Path | None:
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


OGRINFO_PATH = discover_binary(OGRINFO_CANDIDATES)
OGR2OGR_PATH = discover_binary(OGR2OGR_CANDIDATES)


def slugify(value: str) -> str:
    text = unicodedata.normalize("NFD", str(value or ""))
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    text = re.sub(r"[^a-zA-Z0-9]+", "-", text).strip("-").lower()
    return text or "sin-id"


def to_float(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if math.isnan(value):
            return None
        return float(value)
    text = str(value).strip().replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def run_external_command(command: list[str | Path]) -> str:
    prepared = [str(part) for part in command]
    completed = subprocess.run(
        prepared,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="ignore",
        check=False,
    )
    if completed.returncode != 0:
        raise RuntimeError(completed.stderr.strip() or completed.stdout.strip() or "No se pudo ejecutar el comando externo.")
    return completed.stdout


def extract_aprx_entry_names(aprx_path: Path) -> list[str]:
    if not aprx_path.exists():
        return []
    try:
        with zipfile.ZipFile(aprx_path) as archive:
            json_entry = next((entry for entry in archive.namelist() if entry.endswith("GISProject.json")), None)
            if not json_entry:
                return []
            text = archive.read(json_entry).decode("utf-8", errors="ignore")
    except Exception:
        return []
    names = sorted(set(re.findall(r'"name":"([^"]+)"', text)))
    return [name for name in names if name and name not in {"Layers", "Mapa"}]


def extract_ogr_layer_info(gdb_path: Path, layer_name: str) -> dict | None:
    if not OGRINFO_PATH or not gdb_path.exists():
        return None
    try:
        output = run_external_command([OGRINFO_PATH, "-ro", "-so", gdb_path, layer_name])
    except RuntimeError:
        return None

    feature_match = re.search(r"Feature Count:\s*(\d+)", output)
    geometry_match = re.search(r"^Geometry:\s*(.+)$", output, flags=re.MULTILINE)
    extent_match = re.search(
        r"Extent:\s*\(([-\d.]+),\s*([-\d.]+)\)\s*-\s*\(([-\d.]+),\s*([-\d.]+)\)",
        output,
    )
    bbox_utm = None
    geometry = None
    area_ha = None
    if extent_match:
        min_x, min_y, max_x, max_y = (float(extent_match.group(index)) for index in range(1, 5))
        geometry, bbox_utm, area_ha = build_geojson_polygon(
            [
                (min_x, max_y),
                (max_x, max_y),
                (max_x, min_y),
                (min_x, min_y),
            ]
        )
    return {
        "layer": layer_name,
        "featureCount": int(feature_match.group(1)) if feature_match else 0,
        "geometryType": geometry_match.group(1).strip() if geometry_match else "Unknown",
        "bboxUtm": bbox_utm,
        "geometry": geometry,
        "coverageHa": area_ha,
    }


def export_ogr_layer_render_feature(
    gdb_path: Path,
    layer_name: str,
    geometry_type: str = "",
    feature_count: int = 0,
) -> tuple[dict | None, str]:
    if not OGR2OGR_PATH or not gdb_path.exists():
        return None, "none"
    WORK_TMP.mkdir(parents=True, exist_ok=True)
    render_mode = "union"
    sql_geometry = "ST_Union(Shape)"
    if "Polygon" in geometry_type and feature_count > 300:
        render_mode = "envelope"
        sql_geometry = "ST_ConvexHull(ST_Union(Shape))"
    output_path = WORK_TMP / f"{slugify(layer_name)}-{render_mode}.geojson"
    if output_path.exists():
        output_path.unlink()
    try:
        run_external_command(
            [
                OGR2OGR_PATH,
                "-f",
                "GeoJSON",
                output_path,
                gdb_path,
                "-dialect",
                "sqlite",
                "-sql",
                f"SELECT {sql_geometry} AS geometry, COUNT(*) AS feature_count FROM {layer_name}",
                "-t_srs",
                "EPSG:4326",
                "-lco",
                "RFC7946=YES",
                "-lco",
                "COORDINATE_PRECISION=6",
            ]
        )
    except RuntimeError:
        return None, render_mode
    try:
        exported = json.loads(output_path.read_text(encoding="utf-8"))
    except Exception:
        return None, render_mode
    if not isinstance(exported, dict):
        return None, render_mode
    features = exported.get("features", [])
    return (features[0] if features else None), render_mode


def simplify_sensitive_group(label: str) -> str:
    text = slugify(label)
    if "protec" in text:
        return "Protección mínima"
    if "ribere" in text or "cobertura" in text:
        return "Cobertura ribereña"
    if "hidro" in text or "agua" in text or "quebrada" in text or "rio" in text:
        return "Hidrografía sensible"
    if "pend" in text or "geomorf" in text or "talud" in text:
        return "Geomorfología"
    if "model" in text or "multicriterio" in text or "variable" in text:
        return "Modelación multicriterio"
    return label


def utm17s_to_lonlat(easting: float, northing: float) -> tuple[float, float]:
    a = 6378137.0
    ecc_sq = 0.00669437999014
    k0 = 0.9996
    x = easting - 500000.0
    y = northing - 10000000.0
    zone = 17
    long_origin = (zone - 1) * 6 - 180 + 3
    ecc_prime_sq = ecc_sq / (1 - ecc_sq)

    m = y / k0
    mu = m / (
        a
        * (
            1
            - ecc_sq / 4
            - 3 * ecc_sq * ecc_sq / 64
            - 5 * ecc_sq * ecc_sq * ecc_sq / 256
        )
    )

    e1 = (1 - math.sqrt(1 - ecc_sq)) / (1 + math.sqrt(1 - ecc_sq))
    phi1 = (
        mu
        + (3 * e1 / 2 - 27 * e1**3 / 32) * math.sin(2 * mu)
        + (21 * e1**2 / 16 - 55 * e1**4 / 32) * math.sin(4 * mu)
        + (151 * e1**3 / 96) * math.sin(6 * mu)
        + (1097 * e1**4 / 512) * math.sin(8 * mu)
    )

    n1 = a / math.sqrt(1 - ecc_sq * math.sin(phi1) ** 2)
    t1 = math.tan(phi1) ** 2
    c1 = ecc_prime_sq * math.cos(phi1) ** 2
    r1 = a * (1 - ecc_sq) / (1 - ecc_sq * math.sin(phi1) ** 2) ** 1.5
    d = x / (n1 * k0)

    lat = (
        phi1
        - (n1 * math.tan(phi1) / r1)
        * (
            d * d / 2
            - (5 + 3 * t1 + 10 * c1 - 4 * c1 * c1 - 9 * ecc_prime_sq) * d**4 / 24
            + (
                61
                + 90 * t1
                + 298 * c1
                + 45 * t1 * t1
                - 252 * ecc_prime_sq
                - 3 * c1 * c1
            )
            * d**6
            / 720
        )
    )

    lon = (
        d
        - (1 + 2 * t1 + c1) * d**3 / 6
        + (
            5
            - 2 * c1
            + 28 * t1
            - 3 * c1 * c1
            + 8 * ecc_prime_sq
            + 24 * t1 * t1
        )
        * d**5
        / 120
    ) / math.cos(phi1)

    latitude = math.degrees(lat)
    longitude = long_origin + math.degrees(lon)
    return round(longitude, 6), round(latitude, 6)


def build_geojson_polygon(corners_utm: list[tuple[float, float]]) -> tuple[dict, list[float], float]:
    lonlat_ring = [utm17s_to_lonlat(x, y) for x, y in corners_utm]
    if lonlat_ring[0] != lonlat_ring[-1]:
        lonlat_ring.append(lonlat_ring[0])
    xs = [corner[0] for corner in corners_utm]
    ys = [corner[1] for corner in corners_utm]
    bbox = [round(min(xs), 3), round(min(ys), 3), round(max(xs), 3), round(max(ys), 3)]
    area_ha = round(((bbox[2] - bbox[0]) * (bbox[3] - bbox[1])) / 10000, 2)
    return {
        "type": "Polygon",
        "coordinates": [lonlat_ring],
    }, bbox, area_ha


def read_world_file(tfw_path: Path) -> tuple[float, float, float, float, float, float] | None:
    if not tfw_path.exists():
        return None
    lines = [line.strip() for line in tfw_path.read_text(encoding="utf-8", errors="ignore").splitlines() if line.strip()]
    if len(lines) < 6:
        return None
    values = [to_float(line) for line in lines[:6]]
    if not all(value is not None for value in values):
        return None
    return tuple(values)  # type: ignore[return-value]


def corners_from_world_file(image_path: Path, tfw_path: Path) -> tuple[list[tuple[float, float]], list[float], float] | None:
    transform = read_world_file(tfw_path)
    if not transform:
        return None
    with Image.open(image_path) as image:
        width, height = image.size
    a, d, b, e, c, f = transform
    pixel_corners = [
        (-0.5, -0.5),
        (width - 0.5, -0.5),
        (width - 0.5, height - 0.5),
        (-0.5, height - 0.5),
    ]
    corners = []
    for px, py in pixel_corners:
        x = a * px + b * py + c
        y = d * px + e * py + f
        corners.append((x, y))
    geometry, bbox_utm, area_ha = build_geojson_polygon(corners)
    return geometry, bbox_utm, area_ha


def corners_from_geotiff(image_path: Path) -> tuple[dict, list[float], float] | None:
    try:
        with Image.open(image_path) as image:
            tags = image.tag_v2
            scale = tags.get(33550)
            tie = tags.get(33922)
            width, height = image.size
    except Exception:
        return None

    if not scale or not tie or len(scale) < 2 or len(tie) < 6:
        return None

    pixel_x = float(scale[0])
    pixel_y = float(scale[1])
    origin_x = float(tie[3])
    origin_y = float(tie[4])
    corners = [
        (origin_x, origin_y),
        (origin_x + pixel_x * width, origin_y),
        (origin_x + pixel_x * width, origin_y - pixel_y * height),
        (origin_x, origin_y - pixel_y * height),
    ]
    return build_geojson_polygon(corners)


def extract_year(value: str) -> int | None:
    match = re.search(r"(19|20)\d{2}", str(value or ""))
    if not match:
        return None
    return int(match.group(0))


def describe_sector_type(name: str) -> str:
    lowered = name.lower()
    if "r." in lowered or lowered.startswith("r."):
        return "Rio"
    if "uyumbicho" in lowered:
        return "Quebrada / sector lineal"
    if "q." in lowered or lowered.startswith("q."):
        return "Quebrada"
    return "Sector de campo"


def build_sector_features() -> list[dict]:
    field_root = SOURCE_ROOT / "Información de campo"
    features: list[dict] = []
    for folder in sorted(field_root.iterdir()):
        if not folder.is_dir():
            continue

        raster_files = [path for path in folder.rglob("*") if path.is_file() and path.suffix.lower() in {".tif", ".tiff"}]
        orthomosaic = next((path for path in folder.glob("*transparent_mosaic_group1.tif")), None)
        dtm = next((path for path in (folder / "dtm").glob("*_dtm.tif")), None)
        reference = orthomosaic or dtm
        if not reference:
            continue

        tfw = reference.with_suffix(".tfw")
        extent = corners_from_world_file(reference, tfw) or corners_from_geotiff(reference)
        if not extent:
            continue
        geometry, bbox_utm, area_ha = extent
        feature = {
            "type": "Feature",
            "id": slugify(folder.name),
            "geometry": geometry,
            "properties": {
                "name": folder.name,
                "sectorType": describe_sector_type(folder.name),
                "sectorLabel": folder.name.split(".", 1)[-1].strip(),
                "rasterCount": len(raster_files),
                "hasOrthomosaic": orthomosaic is not None,
                "hasDtm": dtm is not None,
                "orthomosaicName": orthomosaic.name if orthomosaic else None,
                "dtmName": dtm.name if dtm else None,
                "orthomosaicMb": round((orthomosaic.stat().st_size / (1024 * 1024)), 1) if orthomosaic else None,
                "dtmMb": round((dtm.stat().st_size / (1024 * 1024)), 1) if dtm else None,
                "coverageHa": area_ha,
                "bboxUtm": bbox_utm,
                "sourceGroup": "informacion_campo",
                "summary": (
                    f"{folder.name}: {len(raster_files)} rasters detectados, "
                    f"{'ortomosaico' if orthomosaic else 'sin ortomosaico'}, "
                    f"{'DTM' if dtm else 'sin DTM'} y cobertura aproximada de {area_ha} ha."
                ),
            },
        }
        features.append(feature)
    return features


def parse_precipitation_station(ws) -> dict:
    monthly = []
    for row in range(9, ws.max_row + 1):
        month_code = ws.cell(row, 1).value
        if month_code not in MONTH_CODES:
            continue
        monthly_mm = to_float(ws.cell(row, 2).value)
        max_24h = to_float(ws.cell(row, 3).value)
        max_day = to_float(ws.cell(row, 4).value)
        rainy_days = to_float(ws.cell(row, 5).value)
        historical = to_float(ws.cell(row, 6).value)
        monthly.append({
            "month": month_code,
            "monthlyMm": monthly_mm,
            "max24hMm": max_24h,
            "maxDay": max_day,
            "rainyDays": rainy_days,
            "historicalMm": historical,
        })

    total_mm = round(sum(item["monthlyMm"] or 0 for item in monthly), 1)
    rainy_days_total = round(sum(item["rainyDays"] or 0 for item in monthly))
    historical_total = round(sum(item["historicalMm"] or 0 for item in monthly), 1)
    max_24h = max((item["max24hMm"] or 0 for item in monthly), default=0)
    wettest = max(monthly, key=lambda item: item["monthlyMm"] or -1, default=None)
    return {
        "kind": "precipitacion",
        "monthlyRecords": monthly,
        "annualMetric": total_mm,
        "annualMetricLabel": f"{total_mm} mm/anio",
        "historicalMetric": historical_total,
        "historicalMetricLabel": f"{historical_total} mm/anio historicos",
        "peakMetric": max_24h,
        "peakMetricLabel": f"{max_24h:.1f} mm max 24h" if max_24h else "Sin maximo 24h",
        "rainyDays": rainy_days_total,
        "headline": wettest["month"] if wettest else "Sin mes pico",
    }


def parse_flow_station(ws) -> dict:
    monthly = []
    for row in range(8, ws.max_row + 1):
        month_code = ws.cell(row, 1).value
        if month_code not in MONTH_CODES:
            continue
        maximum = to_float(ws.cell(row, 2).value)
        minimum = to_float(ws.cell(row, 3).value)
        mean = to_float(ws.cell(row, 4).value)
        historical = to_float(ws.cell(row, 5).value)
        monthly.append({
            "month": month_code,
            "maxLs": maximum,
            "minLs": minimum,
            "meanLs": mean,
            "historicalLs": historical,
        })

    mean_values = [item["meanLs"] for item in monthly if item["meanLs"] is not None]
    historical_values = [item["historicalLs"] for item in monthly if item["historicalLs"] is not None]
    annual_mean = round(statistics.mean(mean_values), 1) if mean_values else 0
    historical_mean = round(statistics.mean(historical_values), 1) if historical_values else 0
    peak_flow = max((item["maxLs"] or 0 for item in monthly), default=0)
    low_flow = min((item["minLs"] for item in monthly if item["minLs"] is not None), default=0)
    return {
        "kind": "caudal",
        "monthlyRecords": monthly,
        "annualMetric": annual_mean,
        "annualMetricLabel": f"{annual_mean} l/s medios",
        "historicalMetric": historical_mean,
        "historicalMetricLabel": f"{historical_mean} l/s historicos",
        "peakMetric": peak_flow,
        "peakMetricLabel": f"{peak_flow:.1f} l/s max" if peak_flow else "Sin caudal maximo",
        "lowMetric": low_flow,
        "headline": "Caudal medio anual",
    }


def build_hydromet_features() -> list[dict]:
    hydro_root = SOURCE_ROOT / "FONAG" / "DATOS HIDROLOGICOS Y METEOROLOGICOS"
    features: list[dict] = []

    for workbook_path in sorted(hydro_root.glob("*.xlsx")):
        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        code = str(worksheet["A1"].value or workbook_path.stem).strip()
        name = str(worksheet["D1"].value or code).strip()
        northing = to_float(worksheet["B3"].value)
        easting = to_float(worksheet["F3"].value)
        altitude = to_float(worksheet["H3"].value)
        if not northing or not easting:
            continue
        longitude, latitude = utm17s_to_lonlat(easting, northing)
        section_label = str(worksheet["A5"].value or "").lower()
        metrics = parse_precipitation_station(worksheet) if "precipit" in section_label else parse_flow_station(worksheet)

        features.append({
            "type": "Feature",
            "id": slugify(code),
            "geometry": {
                "type": "Point",
                "coordinates": [longitude, latitude],
            },
            "properties": {
                "code": code,
                "name": name,
                "stationKind": metrics["kind"],
                "annualMetric": metrics["annualMetric"],
                "annualMetricLabel": metrics["annualMetricLabel"],
                "historicalMetric": metrics["historicalMetric"],
                "historicalMetricLabel": metrics["historicalMetricLabel"],
                "peakMetric": metrics["peakMetric"],
                "peakMetricLabel": metrics["peakMetricLabel"],
                "lowMetric": metrics.get("lowMetric"),
                "rainyDays": metrics.get("rainyDays"),
                "altitudeM": altitude,
                "utmNorthing": northing,
                "utmEasting": easting,
                "headline": metrics["headline"],
                "monthlyRecords": metrics["monthlyRecords"],
                "summary": (
                    f"{name} ({code}) registra {metrics['annualMetricLabel']} "
                    f"frente a {metrics['historicalMetricLabel']}."
                ),
                "sourceGroup": "fonag",
            },
        })
    return features


def build_climate_history() -> list[dict]:
    workbook = load_workbook(SOURCE_ROOT / "INHAMI" / "INHAMI.xlsx", data_only=True, read_only=True)
    records = []
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        annual_totals = []
        month_totals = [0.0] * 12
        month_counts = [0] * 12
        years = []

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            year = row[0]
            if not isinstance(year, (int, float)):
                continue
            year = int(year)
            monthly_values = []
            for index in range(12):
                value = to_float(row[index + 1] if index + 1 < len(row) else None)
                if value is not None:
                    monthly_values.append(value)
                    month_totals[index] += value
                    month_counts[index] += 1
            if monthly_values:
                annual_totals.append(sum(monthly_values))
                years.append(year)

        if not years:
            continue

        month_means = [
            (month_totals[index] / month_counts[index]) if month_counts[index] else 0
            for index in range(12)
        ]
        wettest_index = max(range(12), key=lambda idx: month_means[idx])
        driest_index = min(range(12), key=lambda idx: month_means[idx] if month_counts[idx] else 10**9)
        records.append({
            "stationCode": sheet_name,
            "yearStart": min(years),
            "yearEnd": max(years),
            "sampleCount": len(years),
            "annualMeanMm": round(statistics.mean(annual_totals), 1),
            "wettestMonth": MONTH_NAMES[wettest_index],
            "wettestMonthMm": round(month_means[wettest_index], 1),
            "driestMonth": MONTH_NAMES[driest_index],
            "driestMonthMm": round(month_means[driest_index], 1),
            "summary": (
                f"Serie {sheet_name}: {min(years)}-{max(years)} con media anual "
                f"de {round(statistics.mean(annual_totals), 1)} mm."
            ),
        })
    return sorted(records, key=lambda item: item["annualMeanMm"], reverse=True)


def build_inventory_catalog() -> dict:
    workbook = load_workbook(SOURCE_ROOT / "INFORMACION_HMEJIA.xlsx", data_only=True, read_only=True)
    worksheet = workbook["INFORMACION"]
    categories = []
    current = None

    for row_index in range(9, worksheet.max_row + 1):
        category_id = worksheet.cell(row_index, 1).value
        category_label = worksheet.cell(row_index, 2).value
        item_code = worksheet.cell(row_index, 3).value
        item_label = worksheet.cell(row_index, 4).value

        if category_id is not None or category_label:
            current = {
                "id": str(category_id or len(categories) + 1),
                "label": str(category_label or f"Categoria {len(categories) + 1}").strip(),
                "items": [],
            }
            categories.append(current)

        if current and item_code and item_label:
            current["items"].append({
                "code": str(item_code).strip(),
                "label": str(item_label).strip(),
            })

    return {
        "categories": [
            {
                "id": item["id"],
                "label": item["label"],
                "itemCount": len(item["items"]),
                "items": item["items"],
            }
            for item in categories
        ],
        "categoryCount": len(categories),
        "itemCount": sum(len(item["items"]) for item in categories),
    }


def build_raster_features(folder: Path, group: str) -> list[dict]:
    features: list[dict] = []
    for raster_path in sorted(path for path in folder.rglob("*") if path.is_file() and path.suffix.lower() in {".tif", ".tiff"}):
        if raster_path.name.lower().endswith(".ovr"):
            continue
        extent = corners_from_geotiff(raster_path)
        if not extent:
            tfw_extent = corners_from_world_file(raster_path, raster_path.with_suffix(".tfw"))
            if not tfw_extent:
                continue
            geometry, bbox_utm, area_ha = tfw_extent
        else:
            geometry, bbox_utm, area_ha = extent

        source_kind = "historico"
        name_lower = raster_path.name.lower()
        if "dem" in name_lower:
            source_kind = "dem"
        elif "planet" in name_lower or "nicfi" in name_lower:
            source_kind = "planet"
        elif "orto" in name_lower or "mosaic" in name_lower:
            source_kind = "ortofoto"

        features.append({
            "type": "Feature",
            "id": slugify(f"{group}-{raster_path.stem}"),
            "geometry": geometry,
            "properties": {
                "name": raster_path.stem,
                "sourceGroup": group,
                "sourceKind": source_kind,
                "year": extract_year(raster_path.name),
                "coverageHa": area_ha,
                "bboxUtm": bbox_utm,
                "fileName": raster_path.name,
                "sizeMb": round(raster_path.stat().st_size / (1024 * 1024), 1),
                "summary": f"{raster_path.stem} ({group}) cubre aproximadamente {area_ha} ha.",
            },
        })
    return features


def build_sensitive_consultancy_dataset() -> dict:
    consultancy_workbooks = sorted(path.name for path in CONSULTANCY_ROOT.rglob("*.xlsx"))
    projects = []
    for aprx_path in CONSULTANCY_PROJECTS:
        names = extract_aprx_entry_names(aprx_path)
        highlights = [
            name
            for name in names
            if not any(token in name.lower() for token in ("arcgis", "colors", "provider", "raster calculator", "clip", "merge"))
        ][:10]
        projects.append(
            {
                "name": aprx_path.name,
                "entryCount": len(names),
                "highlights": highlights,
                "summary": (
                    f"{aprx_path.stem} organiza {len(names)} entradas del procedimiento técnico, "
                    f"con énfasis en {', '.join(highlights[:3]) if highlights else 'capas temáticas sensibles'}."
                ),
            }
        )

    sensitive_features: list[dict] = []
    theme_summaries = []
    for theme in CONSULTANCY_SUBPRODUCT_THEMES:
        info = extract_ogr_layer_info(CONSULTANCY_SUBPRODUCT_GDB, theme["layer"]) or {}
        feature_count = int(info.get("featureCount") or 0)
        union_feature, render_mode = export_ogr_layer_render_feature(
            CONSULTANCY_SUBPRODUCT_GDB,
            theme["layer"],
            info.get("geometryType") or "",
            feature_count,
        )
        if not union_feature and not feature_count:
            continue
        if union_feature and union_feature.get("geometry"):
            union_feature["id"] = theme["id"]
            union_feature["properties"] = {
                "name": theme["label"],
                "sensitiveThemeId": theme["id"],
                "sensitiveThemeLabel": theme["label"],
                "sensitiveGroup": theme["group"],
                "sensitiveTone": theme["tone"],
                "sourceLayer": theme["layer"],
                "sourceDatabase": CONSULTANCY_SUBPRODUCT_GDB.name,
                "featureCount": feature_count,
                "renderMode": render_mode,
                "summary": (
                    f"{theme['label']}: {theme['description']} "
                    f"La consultoría consolidó {feature_count} elementos para este tema "
                    f"mediante {'envolvente operativa' if render_mode == 'envelope' else 'geometría integrada'}."
                ),
            }
            sensitive_features.append(union_feature)
        theme_summaries.append(
            {
                "id": theme["id"],
                "label": theme["label"],
                "group": theme["group"],
                "groupLabel": simplify_sensitive_group(theme["group"]),
                "tone": theme["tone"],
                "description": theme["description"],
                "sourceLayer": theme["layer"],
                "featureCount": feature_count or int(info.get("featureCount") or 0),
                "geometryType": info.get("geometryType") or "Unknown",
                "coverageHa": info.get("coverageHa"),
                "bboxUtm": info.get("bboxUtm"),
                "renderMode": render_mode,
            }
        )

    base_layers = []
    for layer in CONSULTANCY_BASE_THEMES:
        info = extract_ogr_layer_info(CONSULTANCY_BASE_GDB, layer["layer"])
        if not info:
            continue
        base_layers.append(
            {
                "layer": layer["layer"],
                "label": layer["label"],
                "group": layer["group"],
                "featureCount": info["featureCount"],
                "geometryType": info["geometryType"],
                "coverageHa": info["coverageHa"],
                "bboxUtm": info["bboxUtm"],
            }
        )

    typology_layers = []
    for layer in CONSULTANCY_TYPOLOGY_THEMES:
        info = extract_ogr_layer_info(CONSULTANCY_TYPOLOGY_GDB, layer["layer"])
        if not info:
            continue
        typology_layers.append(
            {
                "layer": layer["layer"],
                "label": layer["label"],
                "group": layer["group"],
                "featureCount": info["featureCount"],
                "geometryType": info["geometryType"],
            }
        )

    group_counts = Counter(theme["groupLabel"] for theme in theme_summaries)
    polygon_features = sum(
        1 for feature in sensitive_features if feature.get("geometry", {}).get("type", "").endswith("Polygon")
    )
    line_features = len(sensitive_features) - polygon_features

    return {
        "sourceRoot": str(CONSULTANCY_ROOT),
        "summary": {
            "projectCount": len([path for path in CONSULTANCY_PROJECTS if path.exists()]),
            "geodatabaseCount": len([path for path in CONSULTANCY_GDBS if path.exists()]),
            "workbookCount": len(consultancy_workbooks),
            "subproductThemeCount": len(theme_summaries),
            "subproductFeatureCount": sum(theme["featureCount"] for theme in theme_summaries),
            "subproductRenderableThemeCount": len(sensitive_features),
            "baseLayerCount": len(base_layers),
            "baseFeatureCount": sum(item["featureCount"] for item in base_layers),
            "typologyLayerCount": len(typology_layers),
            "typologyFeatureCount": sum(item["featureCount"] for item in typology_layers),
            "polygonFeatureCount": polygon_features,
            "lineFeatureCount": line_features,
            "groupCounts": dict(group_counts),
        },
        "projects": projects,
        "workbooks": consultancy_workbooks,
        "subproductThemes": theme_summaries,
        "baseThemes": base_layers,
        "typologyThemes": typology_layers,
        "subproductCollection": {
            "type": "FeatureCollection",
            "features": sensitive_features,
        },
    }


def build_dataset() -> dict:
    sectors = build_sector_features()
    stations = build_hydromet_features()
    climate_history = build_climate_history()
    inventory = build_inventory_catalog()
    historical = build_raster_features(SOURCE_ROOT / "Cartas Historicas IGM", "cartas_historicas_igm")
    terrain = build_raster_features(SOURCE_ROOT / "MDT historico", "mdt_historico")
    sensitive_consultancy = build_sensitive_consultancy_dataset()

    sector_types = Counter(feature["properties"]["sectorType"] for feature in sectors)
    station_types = Counter(feature["properties"]["stationKind"] for feature in stations)

    return {
        "generatedAt": __import__("datetime").datetime.now(__import__("datetime").UTC).isoformat(timespec="seconds").replace("+00:00", "Z"),
        "sourceRoot": str(SOURCE_ROOT),
        "consultancyRoot": str(CONSULTANCY_ROOT),
        "summary": {
            "fieldSectorCount": len(sectors),
            "hydrometStationCount": len(stations),
            "climateSeriesCount": len(climate_history),
            "historicalRasterCount": len(historical),
            "terrainRasterCount": len(terrain),
            "inventoryCategoryCount": inventory["categoryCount"],
            "inventoryItemCount": inventory["itemCount"],
            "sensitiveConsultancyThemeCount": sensitive_consultancy["summary"]["subproductThemeCount"],
            "sensitiveConsultancyFeatureCount": sensitive_consultancy["summary"]["subproductFeatureCount"],
            "sensitiveConsultancyBaseLayerCount": sensitive_consultancy["summary"]["baseLayerCount"],
            "sensitiveConsultancyProjectCount": sensitive_consultancy["summary"]["projectCount"],
            "sectorTypes": dict(sector_types),
            "stationTypes": dict(station_types),
        },
        "sectors": {
            "type": "FeatureCollection",
            "features": sectors,
        },
        "hydrometStations": {
            "type": "FeatureCollection",
            "features": stations,
        },
        "historicalSources": {
            "type": "FeatureCollection",
            "features": historical,
        },
        "terrainSources": {
            "type": "FeatureCollection",
            "features": terrain,
        },
        "climateHistory": {
            "stations": climate_history,
        },
        "inventoryCatalog": inventory,
        "sensitiveConsultancy": sensitive_consultancy,
    }


def main() -> None:
    dataset = build_dataset()
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(dataset, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"OK {OUTPUT_PATH}")
    print(json.dumps(dataset["summary"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
